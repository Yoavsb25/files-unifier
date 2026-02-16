"""
Excel to PDF converter module.
Handles converting Excel files to PDF format with pagination and wide table support.
"""

from pathlib import Path
from typing import List

from ..core.constants import Constants
from ..core.enums import PageOrientation, PageSize
from ..utils.logging_utils import get_logger

logger = get_logger("pdf_merger.operations.excel_to_pdf_converter")

# Module-level constants
EXCEL_FILE_EXTENSIONS = Constants.EXCEL_FILE_EXTENSIONS
DEFAULT_PAGE_SIZE = Constants.DEFAULT_PAGE_SIZE
DEFAULT_ORIENTATION = Constants.DEFAULT_ORIENTATION
DEFAULT_MAX_COLS_PER_PAGE = Constants.DEFAULT_MAX_COLS_PER_PAGE

__all__ = [
    "convert_excel_to_pdf",
    "_safe_str",
    "_escape_for_paragraph",
    "EXCEL_FILE_EXTENSIONS",
    "DEFAULT_PAGE_SIZE",
    "DEFAULT_ORIENTATION",
    "DEFAULT_MAX_COLS_PER_PAGE",
]


def _safe_str(value) -> str:
    """
    Safely convert a value to string, handling None and other types.

    Args:
        value: Value to convert

    Returns:
        String representation of the value
    """
    if value is None:
        return ""
    return str(value)


def _calculate_column_widths(data: List[List[str]], max_width: float = 3.0) -> List[float]:
    """
    Calculate optimal column widths based on content.

    Args:
        data: Table data (list of rows)
        max_width: Maximum column width in inches (default: 3.0)

    Returns:
        List of column widths in inches
    """
    if not data or not data[0]:
        return []

    num_cols = len(data[0])
    widths = [0.0] * num_cols

    # Find maximum content width for each column
    for row in data:
        for col_idx, cell_value in enumerate(row):
            if col_idx < num_cols:
                # Estimate width based on character count (rough: 10 chars per inch)
                cell_width = min(len(str(cell_value)) / 10.0, max_width)
                widths[col_idx] = max(widths[col_idx], cell_width)

    # Ensure minimum width
    min_width = 0.5
    widths = [max(w, min_width) for w in widths]

    return widths


def _escape_for_paragraph(text: str) -> str:
    """
    Escape XML-special characters so text is safe for ReportLab Paragraph.
    """
    if not text:
        return text
    return text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _split_wide_table(data: List[List[str]], max_cols_per_page: int = 8) -> List[List[List[str]]]:
    """
    Split a wide table into multiple pages.

    Args:
        data: Table data (list of rows)
        max_cols_per_page: Maximum columns per page (default: 8)

    Returns:
        List of table chunks (each chunk is a page)
    """
    if not data:
        return []

    num_cols = len(data[0])

    if num_cols <= max_cols_per_page:
        return [data]

    chunks = []
    for start_col in range(0, num_cols, max_cols_per_page):
        end_col = min(start_col + max_cols_per_page, num_cols)
        chunk = [row[start_col:end_col] for row in data]
        chunks.append(chunk)

    return chunks


def convert_excel_to_pdf(
    excel_path: Path,
    output_path: Path,
    page_size: str = DEFAULT_PAGE_SIZE,
    orientation: str = DEFAULT_ORIENTATION,
    auto_size_columns: bool = True,
    max_cols_per_page: int = DEFAULT_MAX_COLS_PER_PAGE,
) -> bool:
    """
    Convert an Excel file to PDF format using openpyxl and reportlab.
    All worksheets (tabs) are included in a single PDF in workbook order, with each
    sheet starting on a new page. Empty sheets produce a blank page. Supports pagination for
    wide tables and auto-sizing columns.

    Args:
        excel_path: Path to the Excel file (.xlsx or .xls)
        output_path: Path where the PDF will be saved
        page_size: Page size ('letter', 'A4', etc.) - default: 'letter'
        orientation: Page orientation ('portrait', 'landscape') - default: 'portrait'
        auto_size_columns: Whether to auto-size columns based on content - default: True
        max_cols_per_page: Maximum columns per page for wide tables - default: 8

    Returns:
        True if successful, False otherwise

    Raises:
        ImportError: If required libraries are not installed
    """
    if not excel_path.exists():
        logger.error(f"Excel file not found: {excel_path}")
        return False

    if excel_path.suffix.lower() not in EXCEL_FILE_EXTENSIONS:
        logger.error(f"File is not an Excel file: {excel_path}")
        return False

    try:
        # Import required libraries
        try:
            import openpyxl
        except ImportError:
            raise ImportError(
                "openpyxl library is required for Excel file conversion. "
                "Install with: pip install openpyxl"
            )

        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape, letter
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                PageBreak,
                Paragraph,
                SimpleDocTemplate,
                Table,
                TableStyle,
            )
        except ImportError:
            raise ImportError(
                "reportlab library is required for PDF generation. "
                "Install with: pip install reportlab"
            )

        # Determine page size
        page_size_enum = (
            PageSize(page_size.lower())
            if page_size.lower() in [s.value for s in PageSize]
            else PageSize.LETTER
        )
        pagesize_map = {
            PageSize.LETTER.value: letter,
            PageSize.A4.value: A4,
        }
        selected_pagesize = pagesize_map.get(page_size_enum.value, letter)

        # Apply orientation
        orientation_enum = (
            PageOrientation(orientation.lower())
            if orientation.lower() in [o.value for o in PageOrientation]
            else PageOrientation.PORTRAIT
        )
        if orientation_enum == PageOrientation.LANDSCAPE:
            selected_pagesize = landscape(selected_pagesize)

        # Load the Excel workbook
        try:
            wb = openpyxl.load_workbook(str(excel_path), data_only=True)
        except Exception as e:
            logger.error(f"Failed to load Excel file {excel_path.name}: {e}")
            return False

        # Create PDF document once; we will build elements from all worksheets
        doc = SimpleDocTemplate(str(output_path), pagesize=selected_pagesize)
        styles = getSampleStyleSheet()
        cell_style = ParagraphStyle(
            "TableCell",
            parent=styles["Normal"],
            fontSize=9,
            leading=11,
        )
        elements = []
        first_sheet = True

        for sheet in wb.worksheets:
            max_col = sheet.max_column
            max_row = sheet.max_row

            # Page break before each sheet except the first one
            if not first_sheet:
                elements.append(PageBreak())
            first_sheet = False

            if max_row == 0 or max_col == 0:
                # Empty sheet: add a blank page (extra PageBreak yields one empty page)
                elements.append(PageBreak())
                continue

            # Collect all data from this sheet
            data = []
            for row in sheet.iter_rows(
                min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=True
            ):
                row_data = [_safe_str(cell) for cell in row]
                data.append(row_data)

            # Column widths for this sheet
            col_widths = None
            if auto_size_columns:
                col_widths = _calculate_column_widths(data)
                logger.debug(f"Auto-calculated column widths: {col_widths}")

            # Split wide tables into multiple pages
            table_chunks = _split_wide_table(data, max_cols_per_page)
            if len(table_chunks) > 1:
                logger.info(f"Splitting wide table into {len(table_chunks)} pages")

            for chunk_idx, chunk_data in enumerate(table_chunks):
                if chunk_idx > 0:
                    elements.append(PageBreak())

                num_cols = len(chunk_data[0]) if chunk_data else 0
                if col_widths and num_cols:
                    chunk_col_widths = col_widths[:num_cols]
                else:
                    # Equal column widths summing to doc.width when not auto-sizing
                    usable_width_inches = doc.width / inch
                    chunk_col_widths = (
                        [usable_width_inches / num_cols] * num_cols if num_cols else []
                    )

                # Scale column widths so table fits within doc.width (points)
                if chunk_col_widths:
                    total_pts = sum(w * inch for w in chunk_col_widths)
                    scale = min(1.0, doc.width / total_pts) if total_pts > 0 else 1.0
                    col_widths_pts = [w * inch * scale for w in chunk_col_widths]
                else:
                    col_widths_pts = []

                # Convert cell strings to Paragraphs for wrapping; escape XML for Paragraph
                chunk_flowables = [
                    [Paragraph(_escape_for_paragraph(cell), cell_style) for cell in row]
                    for row in chunk_data
                ]

                if col_widths_pts:
                    table = Table(chunk_flowables, colWidths=col_widths_pts)
                else:
                    table = Table(chunk_flowables)

                style = TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 10),
                        ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                        ("TOPPADDING", (0, 0), (-1, 0), 12),
                        ("BACKGROUND", (0, 1), (-1, -1), colors.white),
                        ("TEXTCOLOR", (0, 1), (-1, -1), colors.black),
                        ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
                        ("FONTSIZE", (0, 1), (-1, -1), 9),
                        ("ALIGN", (0, 1), (-1, -1), "LEFT"),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D0D0")),
                        ("LINEBELOW", (0, 0), (-1, 0), 2, colors.HexColor("#2F5597")),
                        (
                            "ROWBACKGROUNDS",
                            (0, 1),
                            (-1, -1),
                            [colors.white, colors.HexColor("#F2F2F2")],
                        ),
                    ]
                )
                table.setStyle(style)
                elements.append(table)

        if not elements:
            logger.warning(
                f"Excel file {excel_path.name} appears to be empty (no data in any sheet)"
            )
            doc.build([])
        else:
            doc.build(elements)

        # Verify the output file was created
        if not output_path.exists():
            logger.error(f"PDF conversion failed: output file not created at {output_path}")
            return False

        sheets_count = sum(1 for s in wb.worksheets if s.max_row > 0 and s.max_column > 0)
        logger.info(f"Successfully converted {sheets_count} sheet(s) from {excel_path.name} to PDF")
        return True

    except ImportError as e:
        logger.error(str(e))
        return False
    except Exception as e:
        logger.error(f"Error converting Excel file {excel_path.name} to PDF: {e}")
        return False
